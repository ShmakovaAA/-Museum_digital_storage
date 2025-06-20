import json
import sqlite3
from openpyxl import load_workbook
from datetime import datetime



# =====================================================
#            Получение и конвертация данных
# =====================================================
def convert_xlsx():

    # Функция для переформатирования даты
    def reformat_date(date_str):
        if not date_str:
            return None
        
        # Преобразуем входное значение в строку
        date_str = str(date_str).strip()
        
        # Если дата уже в формате "yyyy-mm-dd hh:mm:ss", удаляем время
        if date_str.endswith(" 00:00:00"):
            return date_str.split(" ")[0]
        
        # Попробуем преобразовать дату из формата "dd.mm.yyyy"
        try:
            dt = datetime.strptime(date_str, '%d.%m.%Y')
            return dt.strftime('%Y.%m.%d')  # Возвращаем в формате "yyyy.mm.dd"
        except ValueError:
            return date_str  # Если не удалось распарсить, оставляем как есть

    # Функция для обработки диапазонов дат
    def process_dates(date_value):
        if not date_value:
            return {"start_date": None, "end_date": None}
        
        # Преобразуем значение в строку
        date_str = str(date_value).strip()
        
        # Если есть разделитель " - ", разделяем на start_date и end_date
        if ' - ' in date_str:
            start_date, end_date = date_str.split(' - ', 1)
            return {
                "start_date": reformat_date(start_date.strip()),
                "end_date": reformat_date(end_date.strip())
            }
        else:
            return {
                "start_date": reformat_date(date_str),
                "end_date": None
            }

    # Путь к файлу
    file_path = 'data.xlsx'

    # Загружаем Excel-файл
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # Инициализируем список для хранения результатов
    results = []

    # Проходим по каждой строке начиная со второй (первая строка - заголовки)
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Столбец A: Дата выдачи/возврата
        date_info = process_dates(row[0])
        
        # Столбец C: Адрес места экспонирования
        location = row[2] if row[2] else None
        
        # Столбец D: Наименование выставки
        title = row[3] if row[3] else None
        
        # Создаем словарь для текущей строки
        result = {
            "start_date": date_info["start_date"],
            "end_date": date_info["end_date"],
            "title": title,
            "location": location
        }
        
        # Добавляем словарь в результирующий список
        results.append(result)

    # Записываем результаты в JSON-файл
    output_file = 'output.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=4)

    print(f"Результат сохранен в {output_file}")

# =====================================================
#                 Запись данных в БД
# =====================================================
def write_data():
    # Путь к файлу output.json
    json_file_path = 'output.json'

    # Функция для преобразования строки даты в формат YYYY-MM-DD
    def parse_date(date_str):
        if date_str:
            try:
                # Преобразуем строку даты в объект datetime.date
                return datetime.strptime(date_str.replace('-', '.'), '%Y.%m.%d').date().isoformat()
            except ValueError:
                print(f"Неверный формат даты: {date_str}")
                return None
        return None

    # Подключение к базе данных SQLite
    db_path = '../db.sqlite3'
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Создание таблицы Activity, если она не существует
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS Activity (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        title TEXT NOT NULL,
        description TEXT,
        start_date TEXT NOT NULL,
        end_date TEXT,
        event_type TEXT NOT NULL DEFAULT 'exhibition',
        location TEXT,
        related_activities TEXT,
        items TEXT,
        links TEXT,
        preview_photo TEXT
    )
    ''')

    # Чтение данных из JSON-файла
    with open(json_file_path, 'r', encoding='utf-8') as file:
        activities_data = json.load(file)

    # Вставка данных в таблицу Activity
    for activity in activities_data:
        # Извлечение данных из JSON
        start_date = parse_date(activity.get('start_date'))
        end_date = parse_date(activity.get('end_date'))
        title = activity.get('title')
        location = activity.get('location')

        # Подстановка шаблонных значений, если данные отсутствуют
        if not start_date:
            start_date = '1899-01-01'
        if not title:
            title = 'Мероприятие требует редактирования.'
        if not location:
            location = 'Место проведения не указано.'

        # Определение описания (description)
        description = None
        if not activity.get('start_date') or not activity.get('title') or not activity.get('location'):
            description = 'Мероприятие требует редактирования.'

        # Указываем тип мероприятия (event_type)
        event_type = 'exhibition'  # По умолчанию используется значение 'exhibition'

        # SQL-запрос для вставки данных
        cursor.execute('''
        INSERT INTO api_activity (
            title,
            description,
            start_date,
            end_date,
            event_type,
            location
        ) VALUES (?, ?, ?, ?, ?, ?)
        ''', (title, description, start_date, end_date, event_type, location))

    # Сохранение изменений и закрытие соединения
    conn.commit()
    conn.close()

    print("Данные успешно записаны в базу данных.")
   
if __name__ == "__main__":
    convert_xlsx()
    write_data()
    